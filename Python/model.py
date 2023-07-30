# contains ODE solvers
import scipy
# external files containing function vectors for ODE ivp
import Fluxes      as Fl
import Metabolites as M
# for reading in xlsx file configurations
import openpyxl

# Hydraulic model simulator includes 2 fuel sources, 6 tanks (fat, cho, Gr, Gh, ATP, PCR), and vATP that is energy sensitive
def EnergySens():
	# load from excel configuration file
	wb = openpyxl.load_workbook(filename="HydraulicLoading.xlsx")
	ws = wb.active
	xo = 	[ ws.cell(row=3, column=3).value, \
			  ws.cell(row=4, column=3).value, \
			  ws.cell(row=5, column=3).value, \
			  ws.cell(row=6, column=3).value, \
			  ws.cell(row=7, column=3).value, \
			  ws.cell(row=8, column=3).value ]
	param = [ ws.cell(row=3, column=8).value, \
			  ws.cell(row=4, column=8).value, \
			  ws.cell(row=5, column=8).value, \
			  ws.cell(row=6, column=8).value, \
			  ws.cell(row=7, column=8).value, \
			  ws.cell(row=8, column=8).value, \
			  ws.cell(row=9, column=8).value, \
			  ws.cell(row=10, column=8).value, \
			  ws.cell(row=11, column=8).value, \
			  ws.cell(row=12, column=8).value, \
			  ws.cell(row=13, column=8).value, \
			  ws.cell(row=14, column=8).value, \
			  ws.cell(row=15, column=8).value, \
			  ws.cell(row=16, column=8).value, \
			  ws.cell(row=17, column=8).value ]

	# turn off ATP breakdown reaction
	param[14] = 0;
	# make conductance of ATP breakdown reaction small
	param[3] = 0.05*param[3]

	# experiments
	# solver methods really matter here 
	# doc page: https://docs.scipy.org/doc/scipy/reference/generated/scipy.integrate.solve_ivp.html
	# "RK45" (default) becomes unstable after like 3 iterations
	# "RK23" simulates successfully be takes forever because it uses really small step sizes. Tank levels don't change at all
	# "DOP853": unstable after about a dozen iterations
	# "Radau": completes in reasonable time, tank levels change when given sufficient time
	# "BDF": completes in reasonable time, tank levels change when given sufficient time
	# "LSODA": completes in reasonable time, tank levels change given sufficient time
	sol = scipy.integrate.solve_ivp(M.metabolites, [0, 600], xo, args=[param], method="LSODA")

	print(sol.y[-1])
	print(sol.success)

EnergySens()
