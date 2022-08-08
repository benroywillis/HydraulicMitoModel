# contains ODE solvers
import scipy
# external files containing function vectors for ODE ivp
import Fluxes      as Fl
import Metabolites as M
# for reading in xlsx file configurations
import openpyxl

# Hydraulic model simulator includes 2 fuel sources, 6 tanks (fat, cho, Gr, Gh, ATP, PCR), and vATP that is energy sensitive
def EnergySens():

	# load from excel configuration file
	wb = openpyxl.load_workbook(filename="HydraulicLoading.xlsx")
	ws = wb.active
	xo = 	[ ws.cell(row=3, column=3).value, \
			  ws.cell(row=4, column=3).value, \
			  ws.cell(row=5, column=3).value, \
			  ws.cell(row=6, column=3).value, \
			  ws.cell(row=7, column=3).value, \
			  ws.cell(row=8, column=3).value ]
	param = [ ws.cell(row=3, column=8).value, \
			  ws.cell(row=4, column=8).value, \
			  ws.cell(row=5, column=8).value, \
			  ws.cell(row=6, column=8).value, \
			  ws.cell(row=7, column=8).value, \
			  ws.cell(row=8, column=8).value, \
			  ws.cell(row=9, column=8).value, \
			  ws.cell(row=10, column=8).value, \
			  ws.cell(row=11, column=8).value, \
			  ws.cell(row=12, column=8).value, \
			  ws.cell(row=13, column=8).value, \
			  ws.cell(row=14, column=8).value, \
			  ws.cell(row=15, column=8).value, \
			  ws.cell(row=16, column=8).value, \
			  ws.cell(row=17, column=8).value ]

	# turn off ATP breakdown reaction
	param[14] = 0;
	# make conductance of ATP breakdown reaction small
	param[3] = 0.05*param[3]

	# experiments
	sol = scipy.integrate.solve_ivp(M.metabolites, [0, 600], xo, args=[param])

	print(sol.y[0])

EnergySens()
